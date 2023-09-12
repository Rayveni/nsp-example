from fastapi import FastAPI, File, UploadFile, Form
from fastapi.requests import Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse

from io import BytesIO
from pydantic import BaseModel
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook
from tempfile import NamedTemporaryFile

import json
import sys
import traceback
app = FastAPI()


async def catch_exceptions_middleware(request: Request, call_next):
    try:
        return await call_next(request)
    except Exception:
        html_next_line = ' \r\n'
        ex_type, ex_value, ex_traceback = sys.exc_info()
        trace_back = traceback.extract_tb(ex_traceback)

        stack_trace = [f'{html_next_line}Traceback:']
        for trace in trace_back:
            if not('site-packages/starlette/' in trace[0] or 'site-packages/fastapi/' in trace[0]):
                stack_trace.append(
                    f"File : {trace[0]} , Line :{trace[1]}, Func.Name : {trace[2]}, Message :{trace[3]}")
        stack_trace = html_next_line.join(stack_trace)
        return JSONResponse({"message": f"Error:{str(ex_value)}{stack_trace}"}, headers={"status": "false"})


origins = ["*"]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"]
)
app.middleware('http')(catch_exceptions_middleware)


class ReportParams(BaseModel):
    name: str
    age: int


@app.get("/url")
def external_url(request: Request):
    return str(request.url)


@app.post("/upload")
async def upload(report_params: str = Form(...), file: UploadFile = File(...)):
    _report_params = ReportParams.parse_raw(report_params)
    report_params = json.loads(report_params)
    print(report_params)
    wb = Workbook()
    sheet = wb.active
    sheet.title = "My_sheet1"
    sheet2 = wb.create_sheet(title="My_sheet2")
    sheet2.cell(row=2, column=2).value = 2

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        output = BytesIO(tmp.read())
    return StreamingResponse(
        output,
        headers={"Content-Disposition": "attachment",
                 "filename": "data.xlsx", "status": "true"}
    )


@app.get("/xlsx")
async def get_csv_data():

    wb = Workbook()
    sheet = wb.active
    sheet.title = "My_sheet1"
    sheet2 = wb.create_sheet(title="My_sheet2")
    sheet2.cell(row=2, column=2).value = 2

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        output = BytesIO(tmp.read())
    return StreamingResponse(
        output,
        headers={"Content-Disposition": "attachment",
                 "filename": "data.xlsx", "status": "true"}
    )


app.mount("/", StaticFiles(directory="ui", html=True))
